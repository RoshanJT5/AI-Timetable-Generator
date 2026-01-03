"""
CSV Streaming Processor for Large File Uploads
Provides memory-efficient CSV/Excel processing without pandas dependency
"""
import csv
import io
from typing import Iterator, Dict, List, Any
from openpyxl import load_workbook


def process_csv_stream(file_stream, chunk_size: int = 1000) -> Iterator[List[Dict[str, Any]]]:
    """
    Stream CSV file in chunks to avoid loading entire file into memory.
    
    Args:
        file_stream: Binary file-like object (Flask's request.files['file'].stream)
        chunk_size: Number of rows to process per chunk
        
    Yields:
        List of dictionaries representing rows in the current chunk
    """
    # Wrap binary stream in text mode for CSV reader (utf-8-sig handles BOM)
    text_stream = io.TextIOWrapper(file_stream, encoding='utf-8-sig', newline='')
    reader = csv.DictReader(text_stream)
    
    chunk = []
    for i, row in enumerate(reader, 1):
        # Normalize column names to lowercase and strip whitespace
        normalized_row = {k.strip().lower(): v for k, v in row.items() if k}
        chunk.append(normalized_row)
        
        if i % chunk_size == 0:
            yield chunk
            chunk = []
    
    # Yield remaining rows
    if chunk:
        yield chunk


def process_excel_stream(file_stream, chunk_size: int = 1000) -> Iterator[List[Dict[str, Any]]]:
    """
    Stream Excel file in chunks to avoid loading entire file into memory.
    Uses openpyxl's read_only mode for memory efficiency.
    
    Args:
        file_stream: Binary file-like object (Flask's request.files['file'].stream)
        chunk_size: Number of rows to process per chunk
        
    Yields:
        List of dictionaries representing rows in the current chunk
    """
    # Load workbook in read-only mode for memory efficiency
    workbook = load_workbook(file_stream, read_only=True, data_only=True)
    sheet = workbook.active
    
    # Get headers from first row
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter)
    
    # Normalize headers to lowercase and strip whitespace
    headers = [str(h).strip().lower() if h else f'column_{i}' for i, h in enumerate(headers)]
    
    chunk = []
    for i, row_values in enumerate(rows_iter, 1):
        # Create dictionary from headers and values
        row_dict = {}
        for header, value in zip(headers, row_values):
            # Convert None to empty string, everything else to string
            row_dict[header] = '' if value is None else str(value).strip()
        
        chunk.append(row_dict)
        
        if i % chunk_size == 0:
            yield chunk
            chunk = []
    
    # Yield remaining rows
    if chunk:
        yield chunk
    
    workbook.close()


def process_upload_stream(upload_file, chunk_size: int = 1000) -> Iterator[List[Dict[str, Any]]]:
    """
    Automatically detect file type and stream process CSV or Excel files.
    
    Args:
        upload_file: Flask FileStorage object from request.files
        chunk_size: Number of rows to process per chunk
        
    Yields:
        List of dictionaries representing rows in the current chunk
        
    Raises:
        ValueError: If file type is not supported
    """
    filename = upload_file.filename.lower()
    
    if filename.endswith('.csv'):
        yield from process_csv_stream(upload_file.stream, chunk_size)
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        yield from process_excel_stream(upload_file.stream, chunk_size)
    else:
        raise ValueError('Unsupported file type. Upload CSV or Excel (.xlsx, .xls) files only.')


def validate_required_columns(row: Dict[str, Any], required_columns: set) -> bool:
    """
    Check if a row contains all required columns.
    
    Args:
        row: Dictionary representing a single row
        required_columns: Set of required column names (lowercase)
        
    Returns:
        True if all required columns are present, False otherwise
    """
    return required_columns.issubset(set(row.keys()))


def get_missing_columns(available_columns: set, required_columns: set) -> set:
    """
    Get the set of missing required columns.
    
    Args:
        available_columns: Set of available column names
        required_columns: Set of required column names
        
    Returns:
        Set of missing column names
    """
    return required_columns - available_columns


def process_departments_field(departments_value: Any) -> List[str]:
    """
    Process the departments field from CSV/Excel upload.
    Handles comma-separated strings and converts them to a list.
    Normalizes department names to lowercase for consistency.
    
    Args:
        departments_value: Raw value from CSV/Excel (can be string, list, or None)
        
    Returns:
        List of normalized department names (empty list if None or empty string)
        
    Examples:
        >>> process_departments_field("CS, IT, Math")
        ['cs', 'it', 'math']
        
        >>> process_departments_field("Computer Science")
        ['computer science']
        
        >>> process_departments_field("")
        []
        
        >>> process_departments_field(None)
        []
    """
    if departments_value is None or departments_value == '':
        return []
    
    # If already a list, normalize each item
    if isinstance(departments_value, list):
        return [normalize_string(d) for d in departments_value if normalize_string(d)]
    
    # Convert to string and split by comma
    departments_str = str(departments_value).strip()
    
    if not departments_str:
        return []
    
    # Split by comma, normalize each department name
    departments_list = [normalize_string(dept) for dept in departments_str.split(',') if normalize_string(dept)]
    
    return departments_list


def normalize_string(val: Any) -> str:
    """
    Normalize string values for consistent comparison.
    Strips whitespace and converts to lowercase to prevent case-sensitivity issues.
    
    This ensures 'Btech', 'btech', 'BTECH', and ' BTech ' are all treated as 'btech'.
    
    Args:
        val: Input value (can be string, None, or other type)
        
    Returns:
        Normalized lowercase string, or empty string if None/empty
        
    Examples:
        >>> normalize_string("  BTech  ")
        'btech'
        
        >>> normalize_string("Computer Science")
        'computer science'
        
        >>> normalize_string(None)
        ''
        
        >>> normalize_string("")
        ''
    """
    if val is None or val == '':
        return ''
    
    # Convert to string, strip whitespace, and lowercase
    return str(val).strip().lower()
